#!/usr/bin/env python3
"""Parse Futu (富途證券) HK statements into tax-ready CSVs.

Two input formats are auto-detected (no Futu OpenD / API data is ever used):

  1. Annual statement  年度账单 .xlsx   (PREFERRED — cleaner & more accurate)
       Futu app: 我的 → 賬戶詳情 → 年度账单 / annual statement.
       Has 期初/期末 holdings, so cross-year positions (e.g. an option written last
       December and expiring in January) are realized correctly.
  2. Monthly e-statements  月結單 .pdf   (one or many; needs `pdftotext`)
       Futu app: 我的 → 賬戶詳情 → 電子結單 / 月結單.

If a folder contains BOTH, the .xlsx wins. Futu's statements have NO "realized P&L"
section and NO section named "dividend" (dividends hide in 公司行動 / corporate actions),
so realized P&L is reconstructed from the trades here.

Usage:
    python3 parse_futu_statement.py 2025_年度账单.xlsx -o OUTDIR --rate 0.90322
    python3 parse_futu_statement.py /folder -o OUTDIR          # auto: xlsx if present, else pdf
    python3 parse_futu_statement.py a.pdf b.pdf -o OUTDIR

Outputs (CSV, utf-8-sig; YEAR auto-detected):
    futu_<YEAR>_成交明细.csv          stock/option trades; 变动金额 already net of fees
    futu_<YEAR>_股息利息现金流.csv      dividends(公司行动)/interest/deposits/withdrawals/fees
    futu_<YEAR>_期权行权到期.csv        option expiry(EXP) / assignment(ASS) events
    futu_<YEAR>_已实现盈亏_按标的.csv    realized P&L per instrument (average-cost)
    futu_<YEAR>_账户净值.csv           opening/closing balances (cross-check)
    futu_<YEAR>_税务汇总.csv           tax summary: gains/dividends/interest + tax due (--rate)

Requirements: xlsx mode -> `pip install openpyxl`;  pdf mode -> `pdftotext` (poppler).
No personal data is embedded; everything is read from the file(s) you pass in.
"""
from __future__ import annotations
import argparse, csv, glob, os, re, subprocess, sys
from collections import Counter, defaultdict

DIRECTIONS = ("買入開倉", "賣出平倉", "賣出開倉", "買入平倉")        # PDF (traditional)
SKIP_FRAG = ("佣金", "平台使用費", "交易系統使用費", "交收費", "印花稅", "交易費",
             "證監會徵費", "財匯局徵費", "小計", "合計")
# tax-relevant cashflow types (traditional + simplified); excludes 基金申購/贖回 & IPO round-trips
TAX_CASH_TYPES = ("公司行動", "公司行动", "出入金",
                  "證券月度利息扣除", "证券月度利息扣除", "月度利息扣除", "月度利息",
                  "融券利息", "首次稅局登記費", "首次税局登记费", "稅局登記費", "税局登记费")
DIV_TYPES = ("公司行動", "公司行动")
INTEREST_TYPES = ("證券月度利息扣除", "证券月度利息扣除", "月度利息扣除", "月度利息",
                  "融券利息", "首次稅局登記費", "首次税局登记费", "稅局登記費", "税局登记费")
NO_INPUT_HELP = """\
✗ 没找到富途结单（给定路径下没有 .xlsx 或 .pdf）。
富途结单需在 App 内手动导出（无 API / CLI 可自动拉取）：
  • 年度结单（推荐，一个 xlsx）: 富途牛牛 App → 我的 → 账户详情 → 年度账单
  • 月度结单（每月一份 PDF）:    富途牛牛 App → 我的 → 账户详情 → 电子结单/月结单
帮助中心:
  如何获取月结单: https://www.futuhk.com/support/topic2_332
  年度结单同样在手机 App 内获取。
导出后，把文件（或所在目录）作为参数传入，例如:
  python3 parse_futu_statement.py 2025_年度账单.xlsx -o out/ --rate 0.90322
  python3 parse_futu_statement.py /结单所在目录 -o out/ --rate 0.90322"""
FILL_RE = re.compile(
    r"(SEHK|NASDAQ|NYSE|ARCA|AMEX|US)\s+(HKD|USD|CNH|JPY|SGD)\s+"
    r"(\d{4}/\d{2}/\d{2})\s+(\d{4}/\d{2}/\d{2})\s+"
    r"([\d,]+)\s+([\d.]+)\s+([\d,]+\.\d{2})\s+([+-]?[\d,]+\.\d{2})")
CASHFLOW_RE = re.compile(
    r"\s*(\d{4}/\d{2}/\d{2})\s+(增加|減少)\s+(\S+?)\s+(HKD|USD|CNH|JPY|SGD)\s+"
    r"([+-]?[\d,]+\.\d{2})\s*(.*)$")
OPT_EVENT_RE = re.compile(r"Opt (EXP|ASS)-[A-Z]+-([A-Z0-9]+)-\d{8}")
STOCK_CODE_RE = re.compile(r"([0-9]{4,5}|[A-Z]{2,4}\d{6}[CP]\d+)\(([^)]*)\)?")
NUMERIC = re.compile(r"^(HKD|USD|CNH|JPY|SGD|[\d,]+\.?\d*)$")
OPT_CODE_RE = re.compile(r"^([A-Z]{2,4})(\d{6})([CP])(\d+)$")
UNDERLYING = {"MIU": "小米", "TCH": "腾讯", "BAB": "阿里巴巴", "JDC": "京东", "MET": "美团"}


def _f(s): return float(str(s).replace(",", ""))


# PBOC/CFETS RMB central parity (年末中间价) on the tax year's final calendar day.
# Source for 2025: https://www.pbc.gov.cn/zhengcehuobisi/125207/125217/125925/2025123109021714424/index.html
DEFAULT_FX_RATES_BY_YEAR = {
    2025: {
        "HKD": 0.90322,
        "USD": 7.0288,
    },
}


def parse_fx_rates(year, rate, fx_rate_args):
    rates = dict(DEFAULT_FX_RATES_BY_YEAR.get(year, {})) if rate is None and not fx_rate_args else {}
    if rate is not None:
        rates["HKD"] = rate
    for item in fx_rate_args or []:
        if "=" not in item:
            raise argparse.ArgumentTypeError("--fx-rate must be in CCY=RATE format, e.g. USD=7.1")
        ccy, raw_rate = item.split("=", 1)
        ccy = ccy.strip().upper()
        if not ccy:
            raise argparse.ArgumentTypeError("--fx-rate currency cannot be empty")
        try:
            rates[ccy] = float(raw_rate)
        except ValueError as exc:
            raise argparse.ArgumentTypeError(f"--fx-rate has invalid rate: {item}") from exc
    return rates


def amount_to_rmb(value, currency, fx_rates):
    rate = fx_rates.get((currency or "").upper())
    return round(value * rate, 2) if rate is not None else ""


def sums_by_currency(rows, amount_index, currency_index):
    totals = defaultdict(float)
    for row in rows:
        totals[(row[currency_index] or "").upper()] += row[amount_index]
    return dict(totals)


def nice_name(code, fallback=""):
    """Decode an option code to a readable contract name; pass stocks through unchanged."""
    m = OPT_CODE_RE.match(code or "")
    if not m:
        return fallback
    root, date, cp, strike = m.groups()
    return f"{UNDERLYING.get(root, root)} {date} {int(strike) / 1000:.2f} {'沽' if cp == 'P' else '购'}"


# ───────────────────────── PDF parsing ─────────────────────────
def pdf_to_text(path):
    out = subprocess.run(["pdftotext", "-layout", path, "-"], capture_output=True, text=True)
    if out.returncode != 0:
        raise RuntimeError(f"pdftotext failed on {path}: {out.stderr}")
    return out.stdout


def parse_pdf(text):
    """One monthly PDF -> (trades, cashflows, opt_events, opening_nav, closing_nav)."""
    trades, cashflows, opt_raw = [], [], []
    asset_opt = {}
    opening_nav = closing_nav = None
    sec = None
    cur_dir = cur_code = cur_name = ""
    building = False

    for raw in text.splitlines():
        ln, s = raw, raw.strip()
        if "交易-股票和股票期權" in ln:
            sec = "trades"; cur_dir = cur_code = cur_name = ""; building = False; continue
        if "資金進出" in ln:
            sec = "cash"; continue
        if "資產進出" in ln:
            sec = "asset"; continue
        if any(h in ln for h in ("交易-基金", "融資總覽", "重要提示", "期末概覽")):
            if sec == "trades": sec = None
        if "資產淨值" in ln and opening_nav is None:
            nums = re.findall(r"[-\d,]+\.\d{2}", ln)
            if len(nums) >= 2:
                opening_nav, closing_nav = _f(nums[0]), _f(nums[1])

        if sec == "trades":
            hdr = s.lstrip("*").strip()           # '*' = forced-liquidation marker
            is_header = False
            for d in DIRECTIONS:
                if hdr.startswith(d):
                    cur_dir = d
                    rest = hdr[len(d):].split()
                    first = rest[0] if rest else ""
                    cur_code = "" if (not first or NUMERIC.match(first)) else first
                    cur_name = ""; building = bool(cur_code); is_header = True
                    break
            if is_header:
                continue
            m = FILL_RE.search(ln)
            if m:
                mkt, ccy, td, sd, q, p, g, net = m.groups()
                before = ln[:ln.find(mkt)]
                pm = STOCK_CODE_RE.search(before)
                if pm and not building:
                    cur_code, cur_name = pm.group(1), pm.group(2).strip()
                elif "(" in cur_code:
                    cur_code = cur_code.split("(")[0]
                gross, netv = _f(g), _f(net)
                trades.append([td, sd, cur_code, cur_name, cur_dir, ccy,
                               int(_f(q)), float(p), gross, round(abs(abs(netv) - gross), 2), netv])
                building = False
            elif (building and "SEHK" not in ln and s
                  and not any(x in s for x in SKIP_FRAG)):
                if "(" in s:
                    pre, post = s.split("(", 1)
                    cur_code += pre.strip(); cur_name = post.rstrip(")").strip(); building = False
                else:
                    cur_code += s.strip()

        if sec == "asset":
            am = re.search(r"([A-Z]{2,4}\d{6}[CP]\d+)\(([^)]*)", ln)
            if am:
                cm = re.search(r"\+(\d+)\b", ln)
                asset_opt[am.group(1)] = {"name": am.group(2).strip(), "contracts": cm.group(1) if cm else ""}

        if sec == "cash":
            m = CASHFLOW_RE.match(ln)
            if m:
                date, _d, typ, ccy, amt, remark = m.groups()
                cashflows.append([date, typ, ccy, _f(amt), remark.strip()])

        em = OPT_EVENT_RE.search(ln)
        if em:
            dm = re.match(r"\s*(\d{4}/\d{2}/\d{2})", ln)
            nm = re.search(r"\(([^)]*)\)?", ln[ln.find(em.group(2)):]) if "(" in ln else None
            qm = re.search(r"\+(\d+)", ln)
            opt_raw.append({"date": dm.group(1) if dm else "", "code": em.group(2), "kind": em.group(1),
                            "name": nm.group(1).strip() if nm else "", "contracts": qm.group(1) if qm else ""})

    merged = {}
    for e in opt_raw:
        m0 = merged.setdefault((e["code"], e["kind"]),
                               {"date": "", "code": e["code"], "kind": e["kind"], "name": "", "contracts": ""})
        for f in ("date", "name", "contracts"):
            if not m0[f] and e[f]:
                m0[f] = e[f]
    for ev in merged.values():
        a = asset_opt.get(ev["code"])
        if a:
            ev["name"] = ev["name"] or a["name"]; ev["contracts"] = ev["contracts"] or a["contracts"]
    return trades, cashflows, list(merged.values()), opening_nav, closing_nav


def load_pdfs(pdfs):
    trades, cashflows, opt_events, navrows, names = [], [], [], [], {}
    name_re = re.compile(r"([0-9]{4,5}|[A-Z]{2,4}\d{6}[CP]\d+)\(([^)\n]{1,24})\)")
    for pdf in pdfs:
        text = pdf_to_text(pdf)
        for nm in name_re.finditer(text):
            names.setdefault(nm.group(1), nm.group(2).strip())
        tr, cf, ev, onav, cnav = parse_pdf(text)
        trades += tr; cashflows += cf; opt_events += ev
        navrows.append([os.path.basename(pdf), onav, cnav])
    ev_merged = {}
    for e in opt_events:
        ev_merged.setdefault((e["code"], e["kind"]), e)
    nav = (["结单", "期初净值(HKD)", "期末净值(HKD)"], navrows)
    return trades, cashflows, list(ev_merged.values()), {}, nav, names


# ───────────────────────── XLSX parsing (annual statement) ─────────────────────────
def load_xlsx(path, names_hint=None):
    """Annual 年度账单 .xlsx -> trades, cashflows, opt_events, opening_positions, nav, names."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("xlsx mode needs openpyxl — run: pip install openpyxl  (or pass the monthly PDFs instead)")
    wb = openpyxl.load_workbook(path, data_only=True)   # NOT read_only (it misreports dims on these files)

    def sheet(name):
        if name not in wb.sheetnames:
            return {}, []
        rows = [r for r in wb[name].iter_rows(values_only=True) if any(c not in (None, "") for c in r)]
        if not rows:
            return {}, []
        return {c: i for i, c in enumerate(rows[0])}, rows[1:]

    def d8(v):                                          # '20250127' -> '2025/01/27'
        s = str(v)
        return f"{s[:4]}/{s[4:6]}/{s[6:8]}" if len(s) == 8 and s.isdigit() else s

    names = dict(names_hint or {})
    # trades (证券 + 期权 only; funds 基金 are MMF cash-management, excluded like the PDF path)
    ci, rows = sheet("证券-交易流水")
    traw = [r for r in rows if r[ci["品类"]] in ("证券", "股票", "期权")]
    traw.sort(key=lambda r: str(r[ci["成交时间"]]))     # chronological incl. intraday
    trades = []
    for r in traw:
        td = str(r[ci["成交时间"]])[:10].replace("-", "/")
        sd = "" if str(r[ci["交收日期"]]).strip() in ("-", "None", "") else str(r[ci["交收日期"]])
        code = str(r[ci["代码名称"]])
        trades.append([td, sd, code, "", str(r[ci["方向"]]), str(r[ci["币种"]]),
                       int(abs(_f(r[ci["数量/面值"]]))), _f(r[ci["价格"]]),
                       abs(_f(r[ci["成交金额"]])), abs(_f(r[ci["总费用"]])), _f(r[ci["变动金额"]])])

    # cashflows
    ci, rows = sheet("证券-资金进出")
    cashflows = [[d8(r[ci["日期"]]), str(r[ci["类型"]]), str(r[ci["币种"]]),
                  _f(r[ci["变动金额"]]), str(r[ci["备注"]] or "").strip()] for r in rows]

    # option events
    ci, rows = sheet("证券-资产进出")
    opt_events = []
    for r in rows:
        rem = str(r[ci["备注"]] or "")
        kind = "EXP" if "EXP" in rem else "ASS" if "ASS" in rem else ""
        if not kind:
            continue
        opt_events.append({"date": d8(r[ci["日期"]]), "code": str(r[ci["代码名称"]]), "kind": kind,
                           "name": "", "contracts": str(int(_f(r[ci["数量"]])))})

    # opening positions (期初) for the realized ledger; avg = opening cashflow per ABS unit
    ci, rows = sheet("证券-持仓总览")
    opening, hold_mv = {}, defaultdict(float)
    for r in rows:
        code = str(r[ci["代码名称"]]); qty = _f(r[ci["数量/面值"]]); mv = _f(r[ci["市值"]]); ccy = str(r[ci["币种"]])
        per = r[ci["时期类型"]]
        if per in ("期初", "期末"):
            hold_mv[(per, ccy)] += mv
        if per == "期初" and not code.startswith("880") and abs(qty) > 1e-9:
            opening[(ccy.upper(), code)] = (qty, (-mv) / abs(qty))

    # nav cross-check rows: cash (资金总览) + holdings market value, by 期初/期末 × currency
    ci, rows = sheet("证券-资金总览")
    navrows = [[str(r[ci["时期类型"]]), "现金", str(r[ci["币种"]]), round(_f(r[ci["金额"]]), 2)] for r in rows]
    for (per, ccy), mv in sorted(hold_mv.items()):
        navrows.append([per, "持仓市值", ccy, round(mv, 2)])
    navrows.sort(key=lambda x: (x[0] != "期初", x[1], x[2]))
    nav = (["时期类型", "类别", "货币", "金额(原币种)"], navrows)
    return trades, cashflows, opt_events, opening, nav, names


# ───────────────────────── realized P&L (shared) ─────────────────────────
def realized_by_ticker(trades, opt_events, opening=None):
    """Average-cost realized P&L per instrument.

    Open vs close is read from Futu's 開倉/平倉 (开仓/平仓) label; buy vs sell from the
    sign of 变动金额 (net<0 = cash out = buy). This makes it source-agnostic (traditional
    PDF, simplified xlsx, and 强平 forced-liquidation rows all work). `opening` (xlsx only)
    seeds carried-in positions so cross-year options realize correctly.
    Expired/assigned options with no closing trade are closed synthetically at 0.
    Positions still held at year-end keep realized = 0 (unrealized, not taxable this year).
    Each instrument is booked per currency: the same numeric code in HKD vs USD is a
    different position, and realized P&L is never summed across currencies.

    A 平仓/CLOSE with no (or insufficient) prior 开仓 — avg=0, e.g. a position carried from a
    prior year read from monthly PDFs (no 期初 seed), shares transferred in, or a missing
    record — has no real cost basis, so booking it would overstate the gain. Such closes set
    `shorted=True` to be surfaced/handled instead of silently taxed. The xlsx annual path
    avoids most of these by seeding `opening` from 期初持仓.
    """
    book = {}
    ev_codes = {e["code"] for e in opt_events}
    for raw_key, (qty, avg) in (opening or {}).items():
        if isinstance(raw_key, tuple):
            currency, code = raw_key
        else:
            currency, code = "", raw_key
        key = ((currency or "").upper(), str(code))
        book[key] = {"name": "", "currency": key[0], "pos": qty, "avg": avg,
                     "realized": 0.0, "nfills": 0, "sum_net": 0.0, "carry": True, "shorted": False}
    for t in sorted(trades, key=lambda r: r[0]):          # stable -> preserves intraday order
        currency, code, name, d, q, net = (t[5] or "").upper(), t[2], t[3], t[4], float(t[6]), float(t[10])
        key = (currency, code)
        if key not in book and ("", code) in book:        # opening seeded w/o currency -> adopt it
            book[key] = book.pop(("", code))
            book[key]["currency"] = currency
        b = book.setdefault(key, {"name": "", "currency": currency, "pos": 0.0, "avg": 0.0,
                                  "realized": 0.0, "nfills": 0, "sum_net": 0.0, "carry": False, "shorted": False})
        if name and not b["name"]:
            b["name"] = name
        b["nfills"] += 1; b["sum_net"] += net
        buy = net < 0
        if "开仓" in d or "開倉" in d:                     # OPEN (买入开仓=long, 卖出开仓=legit short)
            prev = abs(b["pos"]); tot = b["avg"] * prev + net
            b["pos"] += q if buy else -q
            b["avg"] = tot / abs(b["pos"]) if abs(b["pos"]) > 1e-9 else 0.0
        else:                                             # CLOSE (平仓 / 强平)
            prev = b["pos"]
            b["realized"] += (net / q + b["avg"]) * q
            b["pos"] += q if buy else -q
            if abs(b["pos"]) < 1e-9:
                b["avg"] = 0.0
            if abs(prev) < 1e-9 or q > abs(prev) + 1e-9:  # closing w/o (enough) basis -> phantom gain
                b["shorted"] = True
    for key, b in book.items():
        if abs(b["pos"]) > 1e-9 and key[1] in ev_codes:   # expired/assigned -> close at 0
            b["realized"] += b["avg"] * abs(b["pos"]); b["pos"] = 0.0
    return book


def detect_year(trades, cashflows):
    yrs = Counter(t[0][:4] for t in trades if t[0])
    yrs += Counter(c[0][:4] for c in cashflows if c[0])
    return yrs.most_common(1)[0][0] if yrs else "YEAR"


# ───────────────────────── main ─────────────────────────
def main(argv=None):
    ap = argparse.ArgumentParser(description="Parse Futu HK statements (xlsx or pdf) into tax CSVs")
    ap.add_argument("inputs", nargs="+", help="annual .xlsx, monthly .pdf(s), or a folder")
    ap.add_argument("-o", "--outdir", default="futu_parsed", help="output directory")
    ap.add_argument("--rate", type=float, default=None,
                    help="HKD->RMB year-end 中间价 shorthand, same as --fx-rate HKD=RATE")
    ap.add_argument("--fx-rate", action="append", default=[], metavar="CCY=RATE",
                    help="currency-specific RMB FX rate, e.g. HKD=0.90322 or USD=7.10; may repeat; "
                         "defaults to built-in year-end rates when available")
    ap.add_argument("--on-negative-position", choices=["flag", "exclude", "short"], default="flag",
                    help="how to treat a 平仓 with no/insufficient prior 开仓 (missing cost basis, "
                         "e.g. cross-year positions via PDF, transfers-in): "
                         "'flag' (default) = compute & include in totals but warn; "
                         "'exclude' = drop from totals/tax pending manual review; "
                         "'short' = treat as a confirmed genuine short (compute, include, no warning)")
    args = ap.parse_args(argv)

    xlsxs, pdfs = [], []
    for p in args.inputs:
        if os.path.isdir(p):
            xlsxs += sorted(glob.glob(os.path.join(p, "*.xlsx")))
            pdfs += sorted(glob.glob(os.path.join(p, "*.pdf")))
        elif p.lower().endswith(".xlsx"):
            xlsxs.append(p)
        elif p.lower().endswith(".pdf"):
            pdfs.append(p)
    if not xlsxs and not pdfs:
        sys.exit(NO_INPUT_HELP)
    os.makedirs(args.outdir, exist_ok=True)

    if xlsxs:                                             # xlsx wins (more accurate)
        src = xlsxs[0]
        if len(xlsxs) > 1:
            print(f"note: {len(xlsxs)} xlsx found, using {os.path.basename(src)}")
        if pdfs:
            print(f"note: found annual xlsx — using it (more accurate); ignoring {len(pdfs)} PDF(s) for data")
        # borrow stock names from sibling PDFs if available (best-effort; xlsx has codes only)
        names_hint = {}
        sib = pdfs or sorted(glob.glob(os.path.join(os.path.dirname(src) or ".", "*.pdf")))
        if sib:
            try:
                name_re = re.compile(r"([0-9]{4,5}|[A-Z]{2,4}\d{6}[CP]\d+)\(([^)\n]{1,24})\)")
                for pdf in sib:
                    for nm in name_re.finditer(pdf_to_text(pdf)):
                        names_hint.setdefault(nm.group(1), nm.group(2).strip())
            except Exception:
                pass
        trades, cashflows, opt_events, opening, nav, names = load_xlsx(src, names_hint)
        mode = f"xlsx ({os.path.basename(src)})"
    else:
        trades, cashflows, opt_events, opening, nav, names = load_pdfs(pdfs)
        mode = f"{len(pdfs)} PDF(s)"

    def disp(code, fallback=""):
        # xlsx codes drop leading zeros (9988) vs PDF-harvested names (09988) -> try both
        nm = names.get(code) or names.get(code.zfill(5)) or names.get(code.lstrip("0"))
        return nice_name(code, nm or fallback)

    year = detect_year(trades, cashflows)
    try:
        year_int = int(year)
    except (TypeError, ValueError):
        year_int = None
    try:
        fx_rates = parse_fx_rates(year_int, args.rate, args.fx_rate)
    except argparse.ArgumentTypeError as exc:
        ap.error(str(exc))

    def rmb(v, currency): return amount_to_rmb(v, currency, fx_rates)

    def write(name, header, rows):
        with open(os.path.join(args.outdir, name), "w", newline="", encoding="utf-8-sig") as fp:
            w = csv.writer(fp); w.writerow(header); [w.writerow(r) for r in rows]

    # 1) 成交明细
    tr_sorted = [[r[0], r[1], r[2], disp(r[2], r[3])] + r[4:] for r in sorted(trades)]
    trade_total_rows = [["合计", "", "", "", "", ccy, "", "", "", "", round(total, 2)]
                        for ccy, total in sorted(sums_by_currency(trades, 10, 5).items())]
    write(f"futu_{year}_成交明细.csv",
          ["成交日期", "交收日期", "代码", "名称", "买卖方向", "货币", "数量", "价格", "成交金额", "手续费", "变动金额(净额)"],
          tr_sorted + trade_total_rows)

    # 2) 股息利息现金流
    cf_tax = sorted([c for c in cashflows if c[1] in TAX_CASH_TYPES])
    write(f"futu_{year}_股息利息现金流.csv", ["日期", "类型", "货币", "金额", "备注"], cf_tax)

    # 3) 期权行权到期
    write(f"futu_{year}_期权行权到期.csv", ["日期", "类型", "代码", "名称", "合约数", "说明"],
          [[e["date"], "到期(EXP)" if e["kind"] == "EXP" else "行权(ASS)", e["code"],
            disp(e["code"], e["name"]), e["contracts"],
            "到期作废,保费为已实现收益" if e["kind"] == "EXP" else "被行权->对应正股买入见成交明细"]
           for e in sorted(opt_events, key=lambda e: e["date"])])

    # 4) 已实现盈亏 按标的 (per currency)
    book = realized_by_ticker(trades, opt_events, opening)
    rows, totals = [], defaultdict(float)
    for key in sorted(book, key=lambda c: (c[0], book[c]["realized"])):
        currency, code = key
        b = book[key]; held = abs(b["pos"]) > 1e-9
        if held and b["nfills"] == 0 and b.get("carry"):
            continue                                      # carried-in & untouched & still held -> skip
        realized = round(b["realized"], 2)               # closed-portion gain is realized & taxable
        excluded = b.get("shorted") and args.on_negative_position == "exclude"
        if not excluded:
            totals[currency] += realized                  # even if part of the position is still held
        notes = []
        if held:
            notes.append("年末仍有持仓(空头),未实现不计入" if b["pos"] < 0
                         else "年末仍有持仓,未实现部分不计入(已平仓部分已计入)")
        elif b.get("carry") and b["nfills"] == 0:
            notes.append("含上年结转持仓")
        if b.get("shorted"):
            if args.on_negative_position == "short":
                notes.append("做空/超额平仓(已确认),已计入")
            elif excluded:
                notes.append("⚠ 缺成本基础:已从合计/税务中排除,待核对")
            else:
                notes.append("⚠ 平仓无对应开仓/超额平仓:疑似缺少成本基础(转入/跨年PDF/记录缺失),请核对")
        note = ";".join(notes)
        row = [code, disp(code, b["name"]), currency, b["nfills"], round(b["sum_net"], 2), realized]
        if fx_rates:
            row.append(rmb(realized, currency))
        row.append(note)
        rows.append(row)
    header4 = ["代码", "名称", "货币", "成交笔数", "成交净额合计(原币)", "已实现盈亏(原币)"]
    if fx_rates:
        header4.append("已实现盈亏(RMB)")
    header4.append("备注")
    total_rows = []
    for currency, total in sorted(totals.items()):
        total_rows.append(["合计", "", currency, "", "", round(total, 2)]
                          + ([rmb(total, currency)] if fx_rates else [])
                          + ["已实现合计;年末持有标的不计入"])
    write(f"futu_{year}_已实现盈亏_按标的.csv", header4, rows + total_rows)

    # 5) 账户净值
    write(f"futu_{year}_账户净值.csv", nav[0], nav[1])

    # 6) 税务汇总 (个税 境外所得; this account only — combine 财产转让 across accounts before taxing)
    divs, interests = defaultdict(float), defaultdict(float)
    for c in cashflows:
        currency = (c[2] or "").upper()
        if c[1] in DIV_TYPES and c[3] > 0:               # gross dividends
            divs[currency] += c[3]
        if c[1] in INTEREST_TYPES:                       # paid (negative)
            interests[currency] += c[3]

    th = ["所得项目", "货币", "金额(原币)"] + (["金额(RMB)", "应纳税额(RMB)"] if fx_rates else []) + ["税率", "备注"]
    trows, tax_total_rmb = [], 0.0
    currencies = sorted(set(totals) | set(divs) | set(interests))
    for currency in currencies:
        cap = round(totals.get(currency, 0.0), 2)
        div = round(divs.get(currency, 0.0), 2)
        interest = round(interests.get(currency, 0.0), 2)
        cap_rmb, div_rmb, interest_rmb = rmb(cap, currency), rmb(div, currency), rmb(interest, currency)
        cap_tax = round(cap_rmb * 0.20, 2) if cap_rmb != "" and cap > 0 else (0.0 if cap_rmb != "" else "")
        div_tax = round(div_rmb * 0.20, 2) if div_rmb != "" else ""
        if isinstance(cap_tax, float):
            tax_total_rmb += cap_tax
        if isinstance(div_tax, float):
            tax_total_rmb += div_tax
        trows.append(["财产转让所得·已实现(本账户股票/期权)", currency, cap]
                     + ([cap_rmb, cap_tax] if fx_rates else [])
                     + ["20%", "盈利才计税且需与其他账户同类所得盈亏合并;本表仅本账户,亏损不计税"])
        trows.append(["利息股息红利所得·现金分红(毛额)", currency, div]
                     + ([div_rmb, div_tax] if fx_rates else [])
                     + ["20%", "单独计税,不可扣成本/不可与亏损相抵;境外已预扣可申请抵免"])
        trows.append(["(备查)利息及费用支出", currency, interest]
                     + ([interest_rmb, ""] if fx_rates else [])
                     + ["—", "融资/融券利息、登记费等;非收入,做财产转让可作合理费用参考"])
    if fx_rates:
        trows.append(["合计·本账户应纳税额(估)", "", "", "", round(tax_total_rmb, 2), "",
                      "= 分红税 + 财产转让税(本账户);财产转让最终税额须合并其他账户后确定"])
    else:
        trows.append(["提示", "", "传 --rate <HKD年末中间价> 或 --fx-rate <币种=年末中间价> 可计算人民币与应纳税额"])
    write(f"futu_{year}_税务汇总.csv", th, trows)


    def by_ccy(d): return ", ".join(f"{ccy or 'UNKNOWN'} {v:,.2f}" for ccy, v in sorted(d.items()))
    print(f"source: {mode}  year={year} -> {args.outdir}/")
    print(f"  成交明细:        {len(trades)} 笔, Σ变动金额={by_ccy(sums_by_currency(trades, 10, 5)) or '0.00'}")
    print(f"  股息利息现金流:  {len(cf_tax)} 行, 股息(公司行动+)={by_ccy(divs) or '0.00'}")
    print(f"  期权行权到期:    {len(opt_events)} events")
    print(f"  已实现盈亏:      Σ={by_ccy(totals) or '0.00'}"
          + ("  (含期初结转)" if opening else "  (无期初, PDF口径)"))
    if fx_rates:
        print(f"  税务汇总:        本账户应纳税额估算 RMB {tax_total_rmb:,.2f}")
    shorted = sorted(f"{ccy}:{code}" for (ccy, code), b in book.items() if b.get("shorted"))
    if shorted and args.on_negative_position == "flag":
        print(f"  ⚠ 负持仓提醒(NEGATIVE_POSITION):  {', '.join(shorted)} 出现平仓无对应开仓/超额平仓——"
              f"通常意味着缺少成本基础(跨年持仓走 PDF 口径无期初、转入股票、或记录缺失),已计入合计,可能虚高。\n"
              f"    处理方式:① 改用年度账单 xlsx(含期初持仓)或补上成本基础重跑(最准);"
              f"② --on-negative-position=exclude 先从合计/税务中排除待核对;"
              f"③ 若确实是做空,--on-negative-position=short 确认计入。", file=sys.stderr)
    elif shorted and args.on_negative_position == "exclude":
        print(f"  ⚠ 已排除负持仓标的(NEGATIVE_POSITION):  {', '.join(shorted)} 已从合计/税务中排除,待核对成本基础。",
              file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
